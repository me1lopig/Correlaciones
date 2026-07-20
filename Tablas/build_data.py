"""
build_data.py
=============
Lee el Excel original de tablas de parámetros geotécnicos y genera un YAML
normalizado por cada fuente en ./data/.

Correcciones aplicadas (documentadas en README.md y en cada YAML):
  - Unidades:
      * NAVFAC cohesión  t/m3 -> t/m2   (la cohesión es una tensión)
      * CTE densidades   kN/m2 -> kN/m3  (son pesos específicos)
      * Permeabilidad    m/sg  -> m/s
  - Erratas de etiquetas y nombres de suelo (ver DICC_ERRATAS).

Los VALORES numéricos NO se alteran: solo se limpian de formato (separador de
miles/decimal es-ES), se separan los rangos "a-b" en [min, max] y se convierten
los marcadores "--" y vacíos en null.

Uso:  python build_data.py [ruta_al_excel]
"""
from __future__ import annotations
import sys
import re
from pathlib import Path

import openpyxl
import yaml

AQUI = Path(__file__).resolve().parent
DATA = AQUI / "data"
EXCEL_POR_DEFECTO = "/mnt/user-data/uploads/Tablas_Parametros.xlsx"

# --------------------------------------------------------------------------- #
#  Utilidades de limpieza                                                      #
# --------------------------------------------------------------------------- #
VACIO = {None, "--", "-", "—", ""}

# Erratas de etiquetas / nombres corregidas (traza de saneado)
DICC_ERRATAS = {
    "Arana limpia y mezcla de grava y arena limpia":
        "Arena limpia y mezcla de grava y arena limpia",
    "Arenas biuen graduadas, arenas con grava":
        "Arenas bien graduadas, arenas con grava",
    "Mexcla de limo inorgánico y arcilla":
        "Mezcla de limo inorgánico y arcilla",
    "Rellenso seleccionados compactados":
        "Rellenos seleccionados compactados",
    "Limo poco plastico": "Limo poco plástico",
    "Sedimento muy arcilloso fueretmente orgánico, blando":
        "Sedimento muy arcilloso fuertemente orgánico, blando",
    "Sedimento ligeramente arciloso, orgánico blando":
        "Sedimento ligeramente arcilloso, orgánico blando",
    "Toscos  ": "Toscos",
    "Símbolo de gurpo": "Símbolo USCS",
}


def corrige(texto):
    if isinstance(texto, str):
        texto = DICC_ERRATAS.get(texto, texto).strip()
    return texto


def num_es(valor):
    """Convierte texto es-ES ('1.000', '0,35') a float/int. Devuelve el
    número tal cual si ya lo es, o None si es un marcador de vacío."""
    if valor in VACIO:
        return None
    if isinstance(valor, (int, float)):
        return valor
    s = str(valor).strip()
    if s in VACIO:
        return None
    # separador de miles '.' y decimal ',' (formato español)
    s = s.replace(".", "").replace(",", ".")
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None


def rango(a, b):
    """Devuelve [min, max] a partir de dos celdas, ordenado y con None
    donde no hay dato. Si ambas son vacío -> None."""
    va, vb = num_es(a), num_es(b)
    vals = [v for v in (va, vb) if v is not None]
    if not vals:
        return None
    lo, hi = min(vals), max(vals)
    return lo if lo == hi else [lo, hi]


def rango_texto(s):
    """Parsea 'a-b' o 'a - b' (con formato es-ES) a [min, max]; un único
    número a número; vacío a None."""
    if s in VACIO:
        return None
    s = str(s).strip()
    m = re.match(r"^\s*([\d.,]+)\s*[-–]\s*([\d.,]+)\s*$", s)
    if m:
        return rango(m.group(1), m.group(2))
    return num_es(s)


def perm(s):
    """Parsea permeabilidades en texto tipo '>10^-2', '10^-2 a 10^-5',
    '<10^-9' a [min, max] en m/s (None = extremo abierto)."""
    if s in VACIO:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip().replace(" ", "")
    pot = lambda e: 10.0 ** int(e)
    m = re.match(r"^10\^(-?\d+)a10\^(-?\d+)$", s)
    if m:
        a, b = pot(m.group(1)), pot(m.group(2))
        return [min(a, b), max(a, b)]
    m = re.match(r"^>10\^(-?\d+)$", s)
    if m:
        return [pot(m.group(1)), None]
    m = re.match(r"^<10\^(-?\d+)$", s)
    if m:
        return [None, pot(m.group(1))]
    return None


# --------------------------------------------------------------------------- #
#  Definición de columnas por fuente                                           #
# --------------------------------------------------------------------------- #
def col(campo, etiqueta, unidad, tipo, grupo=None):
    return {"campo": campo, "etiqueta": etiqueta, "unidad": unidad,
            "tipo": tipo, "grupo": grupo}


# --------------------------------------------------------------------------- #
#  Extractores por hoja                                                        #
# --------------------------------------------------------------------------- #
def build_grundbau(wb):
    ws = wb["GrundbauTashenbuch"]
    cols = [
        col("tipo_suelo", "Tipo de suelo", None, "text"),
        col("gran_006", "< 0,06 mm", "%", "text", "Granulometría"),
        col("gran_2",   "< 2,0 mm",  "%", "text", "Granulometría"),
        col("ll", "LL", "%", "rango", "Límites de Atterberg"),
        col("lp", "LP", "%", "rango", "Límites de Atterberg"),
        col("ip", "IP", "%", "rango", "Límites de Atterberg"),
        col("gamma_ap",  "γ aparente",  "kN/m³", "rango", "Peso específico"),
        col("gamma_sum", "γ sumergido", "kN/m³", "rango", "Peso específico"),
        col("humedad", "Humedad natural w", "%", "rango"),
        col("e_def", "Deformabilidad E", "MPa", "rango"),
        col("phi",   "φ'", "°",   "rango", "Resistencia al corte"),
        col("c",     "c'", "kPa", "rango", "Resistencia al corte"),
        col("phi_r", "φ residual", "°", "rango", "Resistencia al corte"),
        col("k", "Permeabilidad k", "m/s", "perm"),   # m/sg -> m/s
    ]
    filas = []

    def par(letra, r):
        return ws[f"{letra}{r}"].value, ws[f"{letra}{r+1}"].value

    def gran(letra, r):
        a, b = par(letra, r)
        if a in VACIO and b in VACIO:
            return None
        if b in VACIO or b is None:
            return str(a).strip()
        return f"{a} – {b}"

    for r in range(4, 38, 2):
        nombre = corrige(ws[f"A{r}"].value)
        if not nombre:
            continue
        filas.append({
            "tipo_suelo": nombre,
            "gran_006": gran("B", r),
            "gran_2":   gran("C", r),
            "ll": rango(*par("D", r)),
            "lp": rango(*par("E", r)),
            "ip": rango(*par("F", r)),
            "gamma_ap":  rango(*par("G", r)),
            "gamma_sum": rango(*par("H", r)),
            "humedad":   rango(*par("I", r)),
            "e_def":     rango(*par("J", r)),
            "phi":       rango(*par("K", r)),
            "c":         rango(*par("L", r)),
            "phi_r":     rango(*par("M", r)),
            "k": (lambda a, b: (None if (a is None and b is None)
                                else [min(x for x in (a, b) if x is not None),
                                      max(x for x in (a, b) if x is not None)]))(
                num_es(ws[f"N{r}"].value), num_es(ws[f"N{r+1}"].value)),
        })
    meta = {
        "id": 1, "fuente_id": "grundbau_taschenbuch",
        "nombre": "Grundbau-Taschenbuch",
        "cita": ("Grundbau-Taschenbuch, 3ª edición, 1ª parte (1980). "
                 "Reproducida en el Curso de Cimentaciones de Rodríguez Ortiz."),
        "nota": ("Cada suelo se define por un rango (valor inferior–superior). "
                 "Valores solo orientativos."),
    }
    return meta, cols, filas


def build_eau(wb):
    ws = wb["EAU_1970"]
    cols = [
        col("tipo_suelo", "Clase de suelo", None, "text"),
        col("gamma_sat", "γ saturado",  "kN/m³", "num", "Peso específico"),
        col("gamma_sum", "γ sumergido", "kN/m³", "num", "Peso específico"),
        col("phi", "φ'", "°", "num"),
        col("c",   "c'", "kPa", "num"),
        col("cu",  "cᵤ", "kPa", "rango"),
        col("e_def", "E", "MPa", "rango"),
    ]
    filas = []
    for r in range(4, 20):
        nombre = corrige(ws[f"A{r}"].value)
        if not nombre:
            continue
        filas.append({
            "tipo_suelo": nombre,
            "gamma_sat": num_es(ws[f"B{r}"].value),
            "gamma_sum": num_es(ws[f"C{r}"].value),
            "phi": num_es(ws[f"D{r}"].value),
            "c":   num_es(ws[f"E{r}"].value),
            "cu":  rango(ws[f"F{r}"].value, ws[f"G{r}"].value),
            "e_def": rango(ws[f"H{r}"].value, ws[f"I{r}"].value),
        })
    meta = {
        "id": 2, "fuente_id": "eau_1970",
        "nombre": "EAU 1970",
        "cita": ("EAU 1970 (Empfehlungen des Arbeitsausschusses "
                 '"Ufereinfassungen" — Comité Alemán de defensa de márgenes). '
                 "Vía Curso de Cimentaciones de Rodríguez Ortiz."),
        "nota": ("cᵤ: cohesión no drenada · φ': fricción drenada · "
                 "c': cohesión drenada · E: módulo de elasticidad."),
    }
    return meta, cols, filas


def build_navfac(wb):
    ws = wb["Suelos Compactados NAVFAC 1971"]
    cols = [
        col("simbolo", "Símbolo USCS", None, "text"),
        col("tipo_suelo", "Clase de suelo", None, "text"),
        col("c_comp", "Cohesión (compactado)", "t/m²", "num"),   # t/m3 -> t/m2
        col("c_sat",  "Cohesión (saturado)",  "t/m²", "num"),
        col("phi", "φ'", "°", "text"),   # incluye valores '>38'
        col("tan_phi", "tan φ", None, "num"),
    ]
    filas = []
    for r in range(4, 20):
        sim = ws[f"A{r}"].value
        if not sim:
            continue
        phi = ws[f"E{r}"].value
        phi = str(phi).strip() if phi not in VACIO else None
        tan = ws[f"F{r}"].value
        filas.append({
            "simbolo": str(sim).strip(),
            "tipo_suelo": corrige(ws[f"B{r}"].value),
            "c_comp": num_es(ws[f"C{r}"].value),
            "c_sat":  num_es(ws[f"D{r}"].value),
            "phi": phi,
            "tan_phi": round(float(tan), 3) if isinstance(tan, (int, float)) else None,
        })
    meta = {
        "id": 3, "fuente_id": "navfac_1971",
        "nombre": "Suelos compactados — NAVFAC 1971",
        "cita": ("Propiedades de suelos compactados, NAVFAC (1971). "
                 "Vía Manual de Taludes del IGME."),
        "nota": ("Clasificación por símbolo USCS. Cohesión en t/m² "
                 "(corregida de t/m³ del original)."),
    }
    return meta, cols, filas


def build_metrosur(wb):
    ws = wb["MetroSur"]
    cols = [
        col("tipo_suelo", "Tipo de suelo", None, "text"),
        col("gamma_ap", "γ aparente", "kN/m³", "num"),
        col("c", "Cohesión c", "kPa", "rango"),
        col("phi", "Ángulo de rozamiento φ", "°", "num"),
        col("e_def", "Módulo de deformación E", "t/m²", "rango"),
        col("poisson", "Coef. de Poisson ν", None, "num"),
        col("balasto", "Coef. de balasto kₕ", "t/m³", "rango"),
    ]
    filas = []
    for r in range(4, 20):
        nombre = corrige(ws[f"A{r}"].value)
        if not nombre:
            continue
        filas.append({
            "tipo_suelo": nombre,
            "gamma_ap": num_es(ws[f"B{r}"].value),
            "c": rango_texto(ws[f"C{r}"].value),
            "phi": num_es(ws[f"D{r}"].value),
            "e_def": rango_texto(ws[f"E{r}"].value),
            "poisson": num_es(ws[f"F{r}"].value),
            "balasto": rango_texto(ws[f"G{r}"].value),
        })
    meta = {
        "id": 4, "fuente_id": "metrosur_1999",
        "nombre": "MetroSur 1999",
        "cita": ("Asignación de parámetros geotécnicos para los proyectos "
                 "de MetroSur (1999). Valores para el diseño de pantallas "
                 "del metro de Madrid."),
        "nota": ("Cuando aparecen dos valores, los autores recomiendan el "
                 "mayor para niveles profundos (>10 m) o con mayor grado de "
                 "consolidación o cementación."),
    }
    return meta, cols, filas


def build_cte_densidades(wb):
    ws = wb["CTE_densidades"]
    cols = [
        col("tipo_suelo", "Tipo de suelo", None, "text"),
        col("peso_sat",  "Peso específico saturado", "kN/m³", "rango"),  # kN/m2 -> kN/m3
        col("peso_seco", "Peso específico seco",     "kN/m³", "rango"),
    ]
    filas = []
    for r in range(4, 8):
        nombre = corrige(ws[f"A{r}"].value)
        if not nombre:
            continue
        filas.append({
            "tipo_suelo": nombre,
            "peso_sat":  rango(ws[f"B{r}"].value, ws[f"C{r}"].value),
            "peso_seco": rango(ws[f"D{r}"].value, ws[f"E{r}"].value),
        })
    meta = {
        "id": 5, "fuente_id": "cte_densidades",
        "nombre": "CTE DB-SE-C — Densidades (D.26)",
        "cita": ("CTE DB-SE-C, tabla D.26. Valores orientativos de densidades."),
        "nota": ("Unidades corregidas a kN/m³ (el original rotulaba kN/m² "
                 "en las columnas 'sup')."),
    }
    return meta, cols, filas


def build_cte_prop(wb):
    ws = wb["CTE_Prop_basicas"]
    cols = [
        col("grupo", "Grupo", None, "text"),
        col("tipo_suelo", "Tipo de suelo", None, "text"),
        col("gamma_ap", "γ aparente", "kN/m³", "rango"),
        col("phi", "Ángulo de rozamiento φ", "°", "rango"),
    ]
    grupo_actual = None
    filas = []
    for r in range(4, 11):
        g = ws[f"A{r}"].value
        if g:
            grupo_actual = str(g).strip()
        nombre = corrige(ws[f"B{r}"].value)
        if not nombre:
            continue
        filas.append({
            "grupo": grupo_actual,
            "tipo_suelo": nombre,
            "gamma_ap": rango_texto(ws[f"C{r}"].value),
            "phi": rango_texto(ws[f"D{r}"].value),
        })
    meta = {
        "id": 6, "fuente_id": "cte_prop_basicas",
        "nombre": "CTE DB-SE-C — Propiedades básicas (D.27)",
        "cita": ("CTE DB-SE-C, tabla D.27. Propiedades básicas de los suelos."),
        "nota": "Valores orientativos.",
    }
    return meta, cols, filas


def build_cte_perm(wb):
    ws = wb["CTE_Permeabilidad"]
    cols = [
        col("tipo_suelo", "Tipo de suelo", None, "text"),
        col("k", "Coeficiente de permeabilidad k_z", "m/s", "perm"),  # m/sg -> m/s
    ]
    filas = []
    for r in range(4, 8):
        nombre = corrige(ws[f"A{r}"].value)
        if not nombre:
            continue
        filas.append({"tipo_suelo": nombre, "k": perm(ws[f"B{r}"].value)})
    meta = {
        "id": 7, "fuente_id": "cte_permeabilidad",
        "nombre": "CTE DB-SE-C — Permeabilidad (D.28)",
        "cita": ("CTE DB-SE-C, tabla D.28. Valores orientativos del "
                 "coeficiente de permeabilidad."),
        "nota": "Rangos en m/s (corregido de 'm/sg' del original).",
    }
    return meta, cols, filas


CONSTRUCTORES = [
    build_grundbau, build_eau, build_navfac, build_metrosur,
    build_cte_densidades, build_cte_prop, build_cte_perm,
]

# --------------------------------------------------------------------------- #
#  Validación de plausibilidad física (falla el build si algo se descuadra)   #
# --------------------------------------------------------------------------- #
LIMITES = {
    "gamma_ap": (8, 24), "gamma_sat": (8, 24), "gamma_sum": (0, 15),
    "gamma_seco": (8, 24), "peso_sat": (8, 24), "peso_seco": (8, 24),
    "phi": (0, 50), "phi_r": (0, 45), "poisson": (0.0, 0.5),
    "k": (1e-13, 1.0),
}


def valida(fuente_id, cols, filas):
    campos = {c["campo"]: c for c in cols}
    for f in filas:
        assert set(f.keys()) == set(campos), (
            f"{fuente_id}: fila con campos != columnas -> {f.get('tipo_suelo')}")
        for campo, v in f.items():
            if campo not in LIMITES or v is None:
                continue
            lo, hi = LIMITES[campo]
            vals = v if isinstance(v, list) else [v]
            for x in vals:
                if isinstance(x, (int, float)):
                    assert lo <= x <= hi, (
                        f"{fuente_id}/{campo}={x} fuera de [{lo},{hi}] "
                        f"({f.get('tipo_suelo')})")


# --------------------------------------------------------------------------- #
#  Main                                                                        #
# --------------------------------------------------------------------------- #
def main():
    ruta = sys.argv[1] if len(sys.argv) > 1 else EXCEL_POR_DEFECTO
    wb = openpyxl.load_workbook(ruta, data_only=True)
    DATA.mkdir(exist_ok=True)

    indice = []
    for construir in CONSTRUCTORES:
        meta, cols, filas = construir(wb)
        valida(meta["fuente_id"], cols, filas)
        doc = {"meta": meta, "columnas": cols, "filas": filas}
        salida = DATA / f"{meta['fuente_id']}.yaml"
        with open(salida, "w", encoding="utf-8") as fh:
            yaml.safe_dump(doc, fh, allow_unicode=True, sort_keys=False,
                           default_flow_style=None, width=100)
        indice.append({"id": meta["id"], "fuente_id": meta["fuente_id"],
                        "nombre": meta["nombre"], "n_filas": len(filas)})
        print(f"  [OK] {meta['fuente_id']:<22} {len(filas):>2} filas -> {salida.name}")

    with open(DATA / "_indice.yaml", "w", encoding="utf-8") as fh:
        yaml.safe_dump({"fuentes": indice}, fh, allow_unicode=True,
                       sort_keys=False)
    print(f"\nGeneradas {len(indice)} fuentes en {DATA}/")


if __name__ == "__main__":
    main()
